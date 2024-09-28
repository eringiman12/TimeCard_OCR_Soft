from genericpath import isfile
import os
import openpyxl
import datetime
from openpyxl import load_workbook
import glob
import sys
import shutil
import tkinter as tk
from tkinter import ttk
import datetime


# 処理パス
Shogo_Puttogether_Folda = sys.argv[1]

def Main(process_year):
   # matching
   Shain_Matching_List = []
   
   # Excelフォルダ
   Excel_Folda = "./asset/etc_files/Excel/"
   # 一時ファイル（各所属）
   tmp_file = Excel_Folda + "/tmp.xlsx"
   
   File_Del(Excel_Folda)
   
   for shozoku_item in glob.glob(Shogo_Puttogether_Folda + "/*/*"):
      # ﾀｲﾑｶｰﾄﾞの処理する為のファイルをローカルにコピーする
      shutil.copy(shozoku_item, tmp_file)
      # 一時ファイルの読み込み
      wb = load_workbook(tmp_file, data_only=True, read_only=True)
   
      # # 商号名と所属リストの作成
      Shogo_name,Shozoku_ary = Shozoku_List_Create(wb)

      # 被り配列を返り値で（名前被り有り）（名前被りなし）
      Cover_Stafdf_ary, Not_Cover_Stafdf_ary = Concat_ary_create(Shozoku_ary)
      
      # 配列の合体（名前被りなし）（名前被り有り）結合
      rt = Kojin_write_ary_create(
          wb, Not_Cover_Stafdf_ary) + Kojin_write_ary_create(wb, Cover_Stafdf_ary)
      Shain_Matching_List.append(rt)
      
      # File_Del(Excel_Folda)
   wb.close()
   
   # 個人表作成（前月からコピペ）
   kojin_path,Now_month_kojin_table_path = Kojin_Table_File_Create(Shogo_name,process_year)

   # # 全値の削除（返り値：月給者リスト,日給社リスト、日リスト、月リスト）
   Employee_ary = Clean_Val_File(kojin_path)
   
   
   # # 従業員配列要素詳細[名前,有休,残業・遅刻早退,出勤数,合計出勤時間,所定出勤日数,合計昼休憩時間]
   # # 集計内容のs記載
   write_syukei(Shain_Matching_List, Employee_ary,kojin_path)

def write_syukei(Syukei_ary, Employee_ary,kojin_path):
   # シート読み込み
   wb = load_workbook(kojin_path, keep_vba=True)

   # print(Syukei_ary)
   # 集計結果配列ループ
   for syukei_item in Syukei_ary:
      for kojin_item in syukei_item:
         Employee_Name = kojin_item[0]
         
         # 一致ループ
         for Employee_val in Employee_ary:
            # 従業員名（比較）
            Employee_Name_Compare = Employee_val[1]
            # Compare
            # 比較用配列と処理用配列の従業員名を比較
            if Employee_Name_Compare == Employee_Name:
               # 列名
               col = Employee_val[2]
               # 処理シート名
               sheet = wb[Employee_val[0]]
               
               sheet.cell(row=33, column=col).value = kojin_item[3]
               
               # 出勤数を記載する
               sheet.cell(row=34, column=col).value = kojin_item[3]
               # 有休日数を記載する
               sheet.cell(row=35, column=col).value = kojin_item[1]
               
               # 有休時間
               Paid_time_Total = kojin_item[1] * 8
               # 所定労働時間
               Simple_working_hours = kojin_item[5] * 8
               # 休日欠勤数
               Holiday_absent_cnt = (kojin_item[3] - kojin_item[5]) * 8
               
               # 残業時間のトータル
               Total = kojin_item[4] + Paid_time_Total - Simple_working_hours - Holiday_absent_cnt + kojin_item[6]
               
               # 休日出勤・欠勤
               Holiday_work_absenteecnt = kojin_item[3] + kojin_item[1] - kojin_item[5] 
               
               # 月給日給シートで記載行が違う
               # 正なら残業時間、負なら遅刻
               if Total > 0:
                  # 残業時間に記載する
                  sheet.cell(row=39, column=col).value = Total
               else:
                  # 遅刻早退行に記載する
                  sheet.cell(row=45, column=col).value = Total        
               
                  
               # 正なら残業時間、負なら遅刻  
               if Holiday_work_absenteecnt > 0:
                  # 休日出勤に記載する
                  sheet.cell(row=37, column=col).value = Holiday_work_absenteecnt
               else:
                  # 欠勤に記載する
                  sheet.cell(row=36, column=col).value = Holiday_work_absenteecnt
               break
         
   wb.save(kojin_path)
   

def Clean_Val_File(path):
   wb = load_workbook(path,keep_vba=True)
   # 月給者の配列
   monthly_salary_ary = []
   # 社保日給の配列
   shaho_nikyu_ary = []
   # 日給配列  
   Daily_wage_ary =[]
   # 月給配列
   Monthly_payment_ary = []
   
   # 名前行
   Name_row = 30
   # 開始列
   st_col = 4
   # セルの値削除開始行
   sr_row = 33
   # セルの値削除終了行
   ed_row = 45
   
   for sheet in wb.worksheets:
      # シート名から空白の削除
      r = sheet.title.replace(" ", "").replace("　", "")
      
      if r == "社保月給":
         # シートオブジェクト、名前行、開始列、開始行、終了行
        monthly_salary_ary =  Clean_val(sheet, Name_row, st_col, sr_row, ed_row)
      elif r == "社保日給":
        shaho_nikyu_ary = Clean_val(sheet, Name_row, st_col, sr_row, ed_row)
      elif r == "日月":
        # 日給配列に格納   
        Daily_wage_ary = Clean_val(sheet, Name_row, st_col, sr_row, ed_row)
        # 月給処理開始列の特定
        Month_Col = len(Daily_wage_ary) + st_col + 1
        # 月給配列の格納   
        Monthly_payment_ary = Clean_val(sheet, Name_row, Month_Col, sr_row, ed_row)
        
   wb.save(path)
   wb.close()
   
   rtn_ary = monthly_salary_ary + shaho_nikyu_ary + Daily_wage_ary +Monthly_payment_ary
   return rtn_ary


# シート名、名前行、開始列、開始行、終了行
def Clean_val(sheet, name_row, name_col, strow, ed_row):
   # 返り値配列
   rtn_ary = []
   while True:
      staff_ary = []
      # 名前取得
      val = sheet.cell(row=name_row, column=name_col).value
      
      if val == "" or val == None or val=="日給小計" or val=="月給小計":
         break
      else:
         for row in range(strow, ed_row + 1):
            sheet.cell(row=row, column=name_col).value = ""
      
      # if val != "" and val !=None:
      # 文字列ならtrue
      if val.isdecimal() == False:             
         val = val.replace(" ","").replace("　","")

      staff_ary.append(sheet.title)
      staff_ary.append(val)
      staff_ary.append(name_col)
      rtn_ary.append(staff_ary)
          
      name_col += 1
   
   # 配列詳細（シート名,名前,列）
   return rtn_ary

# 所属先リストの作成
def Shozoku_List_Create(wb):
   # 所属配列の初期化
    Shozoku_ary = []

    # 処理商号名初期化
    Shogo_name = ""
    
    for sheet in wb.worksheets:
        kakuno_ary = []
        # シートの読み込み
        ws = wb[sheet.title]
        
        # 社員リスト意外を処理
        if sheet.title != "社員リスト" and sheet.title != "社員現場詳細リスト":
            # 名前
            Name_Get = ws.cell(row=1, column=5).value
            
            if Name_Get == None:
                Name_Get = ws.cell(row=2, column=5).value
            
            if Name_Get == None:
                   Name_Get = ws.cell(row=3, column=5).value

            # 副現場
            # sub_site = ws.cell(row=2, column=3).value
            if Name_Get == "" or Name_Get == None:
               Name_Get = ws.cell(row=2, column=5).value

               if Name_Get == "" or Name_Get == None:
                  Name_Get = ""

            #　空白削除
            RepName_Get = Name_Get.replace("　", "").replace(" ", "")

            # 主現場要素格納
            kakuno_ary.append(RepName_Get)
            # シート名格納
            kakuno_ary.append(sheet.title)

            # 配列の探索
            Shozoku_ary.append(kakuno_ary)
        else:
           if sheet.title == "社員リスト":
                  
               # 社員リストの商号名を取得
               Shogo_name = ws.cell(row=1, column=1).value
    
    return Shogo_name,Shozoku_ary


# 個人表作成
def Kojin_Table_File_Create(Shogo_name,get_yymm):
   
   if "特定の会社" in Shogo_name:
      Shogo_name = Shogo_name.replace("株式会社", "")
      
      R_ari_year  = get_yymm[:get_yymm.find(".")]
      month = int(get_yymm[get_yymm.find(".") + 1:])
      R_Nasi = get_yymm[get_yymm.find("R") + 1:]
      R_Nasi_year = int(R_Nasi[:get_yymm.find(".") - 1])
      
      Last_month_kojin_table_path = "共有フォルダ\\●" + Shogo_name + "　　給与●\\" + R_ari_year + "給与\\" + \
          get_yymm + "\\出向\\" + R_Nasi + "月会社名（社員番号順）.xlsm"

      if month != 12:
         month += 1
      else:
         R_Nasi_year += 1
         month = 1
      
      
      Now_month_kojin_table_path = "フォルダ" + Shogo_name + "　　給与●\\R" + str(R_Nasi_year) + "給与\\R" + \
          str(R_Nasi_year) + "." + str(month) + "(準備)\\出向\\" + str(R_Nasi_year) + \
          "." + str(month) + "月会社名（社員番号順）.xlsm"
      
      Now_month_Pearsor_Tenki_Path = ".\\asset\\etc_files\\Excel\\"+ str(R_Nasi_year) + "." + str(month)+ "月個人票　会社名 出向（社員番号順）.xlsm"
      # 個人表が無ければ作成
      if not os.path.isfile(Now_month_Pearsor_Tenki_Path):
         shutil.copy(Last_month_kojin_table_path, Now_month_Pearsor_Tenki_Path)
      
      return Now_month_Pearsor_Tenki_Path,Now_month_kojin_table_path
        
def Kojin_write_ary_create(wb, Sheet_ary):
   #  wb = load_workbook(f_path)
    af = ""
    # 有休日数変数の初期化  
    yukyu = 0
    # 残業遅刻変数
    o_time_be_time = 0
    # 出勤日数カウント配列
    Syu_kin_total_day =[]  
    # 出勤時間
    Syu_kin_time_totalling = 0
    # 昼休憩日数 
    Lunch_break_days = 0
    
    #  名前、稼働時間に二次元配列
    kojin_table_write_ary = []

    for item in Sheet_ary:
        i = 1
        Fi = ""
        # 最終ループ回数  
        lat_len = len(item)
        # 名前の初期化   
        Name = ""
        for item2 in item:
            
            if i % 2 == 0:
               if i == 2:
                  Fi = item2
                  
               if Fi != "":
                  # 貼り付け元ブック
                  Pasted_ws = wb[item2]
                  # 有休日数の取得
                  tmp_yukyu = Sheet_value_Get(Pasted_ws,3,8)
                  # 残業遅刻の取得
                  tmp_o_time_be_time = Sheet_value_Get(Pasted_ws, 4, 4)
                  # 指定出勤日数
                  tmp_o_time_totalling = Sheet_value_Get(Pasted_ws, 6, 14)
                  # 昼休憩日数
                  tmp_Lunch_break = Sheet_value_Get(Pasted_ws, 4, 5)
                  
                  # 以前の処理が空でなく違う場合にtrue
                  if af != "" and af != Fi:
                      # 有休変数を0にする 
                      yukyu = 0
                      # 残業遅刻
                      o_time_be_time = 0
                      # 出勤数
                      Syu_kin_total_day = []
                      # 出勤時間  
                      Syu_kin_time_totalling = 0
                      # 昼休憩日数の初期化
                      Lunch_break_days = 0  
                  
                  # 出勤日のカウント
                  Syu_kin_total_day += Syukkin_cnt_ary(Pasted_ws, 4,14)
                  # 稼働時間の加算
                  yukyu += int(tmp_yukyu)
                  # 残業遅刻の加算
                  o_time_be_time += tmp_o_time_be_time
                  
                  # 出勤時間の加算
                  Syu_kin_time_totalling += tmp_o_time_totalling
                  # 昼休憩日数の加算
                  Lunch_break_days += tmp_Lunch_break
                  

                  # 最終加算された値を転記する。
                  if lat_len == i:
                     # 一次元一時配列
                     rtn_ary = []
                     # 名前の格納
                     rtn_ary.append(Name)
                     # 稼働時刻の格納
                     rtn_ary.append(yukyu)
                     # 残業遅刻の格納
                     rtn_ary.append(o_time_be_time)
                     # 重複被りを削除し出勤日数を割り出す
                     rtn_ary.append(len(list(set(Syu_kin_total_day))))
                     # 出勤時間の合計
                     rtn_ary.append(Syu_kin_time_totalling)
                     # 所定出勤数
                     rtn_ary.append(Sheet_value_Get(Pasted_ws, 3, 7))
                     # 昼休憩日数
                     # 休憩時間
                     rtn_ary.append(Lunch_break_days)
                     # 個人表に書き込む形に成型した2次元配列
                     kojin_table_write_ary.append(rtn_ary)
                  
                  # 以前の処理シートタイトル
                  af = Fi
            else:
               Name = item2
            i += 1
    

    return kojin_table_write_ary

def Sheet_value_Get(sheet,col,genzan):

   # last_row = Last_Row_Search(sheet,4,1)
   
   last_row = sheet.max_row - genzan
   # 稼働時間
   rt_val = sheet.cell(row=last_row, column=col).value

   return rt_val 

def Syukkin_cnt_ary(sheet, st_row, genzan):
   last_row = sheet.max_row - genzan
   # 返り値配列
   rtn_ary = []
   
   # 行ループ
   for row in sheet.iter_rows(min_row=st_row, max_row=last_row, max_col=2):
      if row[1].value != None and row[1].value != "":
         rtn_ary.append(row[0].value)
      
   return rtn_ary
   
def Concat_ary_create(data):
   # 名前検索配列に代入
   Shozoku_serch_ary = data
   # 名前
   Cover_Staff_ary = []
   Not_Cover_Staff_ary = []
   
   # 被り配列の作成
   for item in data:
      # 被りカウント
      cnt = 0
      # 名前変数
      Name = item[0]
      # 名前リスト
      Name_list = []
      # 所属名配列ループ
      for item2 in Shozoku_serch_ary:
         # 名前被りの検索
         if item2[0] == Name:
            # 名前格納
            Name_list.append(item2[0])
            # シート名格納
            Name_list.append(item2[1])
            cnt += 1

      # 名前被りが存在する場合にtrue
      if cnt != 1:
         # 存在するかどうか
         sonzai = True
         for item in Cover_Staff_ary:
             if item[0] == Name:
                sonzai = False
         # 名前被り配列に追加
         if sonzai:
            Cover_Staff_ary.append(Name_list)
      else:
         # 名前被り以外
         Not_Cover_Staff_ary.append(Name_list)
               
   return Cover_Staff_ary, Not_Cover_Staff_ary

def File_Del(path):
   # if os.path.isfile(path):
   shutil.rmtree(path)
   os.mkdir(path)
         
if __name__ == "__main__": 
   now = datetime.datetime.now()
   wareki_year = now.year % 100 - 18
   wareki_alp = "R"
   now_month = now.month - 2 
   # now_month = 10
   
   shoki = 1
   Wareki_Year_ary = []
   for i in range(0,12):
      if int(now_month) + i <= 12:
         # print(wareki_alp+str(wareki_year)+ "." +str(int(now_month) + i))
         Wareki_Year_ary.append(wareki_alp+str(wareki_year)+ "." +str(int(now_month) + i))
      else:
         # print(wareki_alp+str(wareki_year + 1)+ "." +str(int(shoki)))
         Wareki_Year_ary.append(wareki_alp+str(wareki_year + 1)+ "." +str(int(shoki)))
         shoki += 1
        
         
   # ルートフレームの作成
   root = tk.Tk()
   root.geometry("300x100")
   # コンボボックスの作成(rootに配置,リストの値を編集不可(readonly)に設定)
   combo = ttk.Combobox(root, state='readonly')
   # リストの値を設定
   combo["values"] = (Wareki_Year_ary)
   # デフォルトの値を食費(index=0)に設定
   combo.current(0)
   # コンボボックスの配置
   combo.pack()

   # ボタンの作成（コールバックコマンドには、コンボボックスの値を取得しprintする処理を定義）
   button = tk.Button(text="処理開始",command=lambda:Main(combo.get()))
   # ボタンの配置
   button.pack()

   root.mainloop()


   # Main("w")

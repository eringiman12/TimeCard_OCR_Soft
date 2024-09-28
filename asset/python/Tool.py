from genericpath import isfile
import glob
from importlib.resources import path
from pickle import FALSE
import shutil
import os
import cv2
import numpy as np
from PIL import Image as Match_Image
import io

import re
from string import digits

import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.chart.axis import Scaling
from openpyxl.chart import BarChart, Reference, Series, LineChart
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, fills
from openpyxl.workbook.defined_name import DefinedName
from openpyxl import utils

from natsort import natsorted

import jpholiday
from datetime import date
import datetime
from datetime import timedelta
import calendar
from openpyxl.drawing.image import Image

import numpy as np
import random

from google.cloud import vision

os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = r"vision api key のファイルパス"
client = vision.ImageAnnotatorClient()

############################################################################ 
# 各メソッドで便利なツールを使えるクラスオブジェクト
# ----Folda_Createメソッド-----
# フォルダを作成する
# 返り値は作成したフォルダのパス
# 引数処理したい親フォルダ

# ----TimeCard_Folda_Createメソッド-----
# ﾀｲﾑｶｰﾄﾞの所属毎に個人でフォルダを分ける
# ﾀｲﾑｶｰﾄﾞが奇数毎時は、エラーパスを返り値にし、致命的なエラーとしてエラーフォルダのパスを返す
# 途中でとまった場合は、画像の枚数を確認すること。

# ----Time_Parts_2_Ch_Imgメソッド-----
# 画像の指定領域を二値化した画像を生成する
# 引数
# imgpath:画像ファイルのパス
# expdirpath:出力先フォルダのパス 
# x,y:二値化の開始点のx,y座標
# x2,y2:二値化の終了点のx,y座標
############################################################################

# 便利ツール
class Tool_object():
   # フォルダ作成関数（返り値：作成先フォルダパス）
   def Folda_Create(self,Path):
       # ﾀｲﾑｶｰﾄﾞ晴らしフォルダがないとき  
       if os.path.isdir(Path) == False:
            os.makedirs(Path)
       return Path

   def detect_text(self, path):
      # 返り値変数の取得  
      rtn_text = ""
      
      # 画像の読み込み
      with io.open(path, 'rb') as image_file:
         content = image_file.read()
      # vision APIの呼び出し
      image = vision.Image(content=content)

      # レスポンス帰り井
      response = client.text_detection(image=image)
      # テキスト配列を返す
      texts = response.text_annotations
      
      for item in texts:
         rtn_text = item.description
         
      rtn_text = rtn_text.replace(':', '')
      

      return rtn_text
   
   # 時刻のランダム作成
   def Randum_Time_crete(self):
      hour_1 = random.randint(0,2)
      hour_2 = random.randint(0,9)
      
      minute_1 = random.randint(0,5)
      minute_2 = random.randint(0,9)
      
      str_time = ""
      if hour_2 <= 2:
         str_time = str(hour_1) + str(hour_2) + str(minute_1) + str(minute_2)
      else:
         str_time = str(hour_2) + str(minute_1) + str(minute_2)
         
      return int(str_time)

   def Ocr_Process(self,Time_Foda,Time_Card_type):
      Ocr_Result_ary = []
      tmp_ary = []

      i = 1
      #   OCR処理：記号は「：」以外除去
      for row_fol in Time_Foda :
         
         # 行カウント取得
         row = os.path.basename(os.path.dirname(row_fol))
         # 列カウント取得
         col = os.path.basename(row_fol).replace('.jpg', '')
   #       print(row_fol)
         
         # # # OCR取得
         # ocr_detect = re.sub('[^0-9:\n]', "", self.detect_text(row_fol))
         
         # # OCRで取得した値が空でなければ数値変換
         # if ocr_detect != "":
         #    # 数値にキャスト
         #    ocr_detect = int(ocr_detect)

         #    # 値が2桁以下の場合は削除 
         #    if len(str(ocr_detect)) <= 2:
         #       ocr_detect = ""
         
         # ocr_detect = "行：" + str(row) + "列" + str(col)
         ocr_detect = self.Randum_Time_crete()
         # 一次元配列に格納
         tmp_ary.append(ocr_detect)      
         
         # 4列眼処理時に二次元配列に格納
         if i == 4:
            if Time_Card_type == "Mx":
               if int(row) != 17:
                  Ocr_Result_ary.append(tmp_ary)
            tmp_ary = []
            i = 1
         else:
            i+= 1
      return Ocr_Result_ary
   
   def Ary_Concat(self,Yuan_ary,First_ary):
      # インデックス番号の初期化
      l = 0
      # 削除配列
      del_ary = []
      
      # 結合先に結合元の配列の要素を結合
      for item in Yuan_ary:
         # 結合先の値が空でなければTRUE
         if First_ary[l] != "":
            # 結合先に結合元の値を挿入
            item.insert(0, First_ary[l])
         else:
            # 空インデックスを挿入
            del_ary.append(l)
         l += 1
         
      return del_ary
      
   def Txt_Read(self,path,get_in):
      # ﾀｲﾑｶｰﾄﾞの種類と行列の入った配列
      TimeCard_rc_syurui= open(path, 'r')
      # ﾀｲﾑｶｰﾄﾞ情報の読み取り   
      TimeCard_rc_syurui_ary = TimeCard_rc_syurui.read().split('\n')[get_in]
      # ﾀｲﾑｶｰﾄﾞテキスト閉じる   
      TimeCard_rc_syurui.close()
      
      return TimeCard_rc_syurui_ary
      
   def TimeCard_Folda_Create(self,path):
      # 個人分け時に奇数でエラーを返す(返り値：真偽値,エラーパス（エラー発生時）)
      # ファイル数ジャッジ
      judge = True
      # 個人仕分けフォルダパス一覧
      Kojin_Create_Fopath = glob.glob(path + "/*")

      # ファイル数ジャッジ
      judge = True
      # エラー発生パス
      Error_Path = ""
      
      # 指定フォルダの配下
      for Shogo in Kojin_Create_Fopath:
         # 商号フォルダ内のリスト一段作成 
         files = os.listdir(Shogo)
         files_file = [f for f in files if os.path.isfile(os.path.join(Shogo, f))]
         files_dir = [f for f in files if os.path.isdir(os.path.join(Shogo, f))]
         
         File_ary = []
         for item in files_file:
            if "Thumbs.db" not in item:
               File_ary.append(item)
         
         
         # ファイル数が偶数の場合のみ処理  
         if len(File_ary) % 2 == 0:
            #  ナンバーフォルダ配列
            numfol = []
            
            # ファイル数に応じてフォルダ作成
            for cnt in range(len(files_dir),int(len(files_dir)) + int(len(File_ary) / 2)):
               # フォルダ作成
               self.Folda_Create(Shogo + "/" + str(cnt + 1))  
               # 配列に格納
               numfol.append(Shogo + "/" + str(cnt + 1))
            
            # インデックス変数
            i = 0
            # 番号フォルダループ
            for item in numfol:
               # ナンバーフォルダ野中が空の場合異動
               if len(glob.glob(item+"/*")) == 0:
                  shutil.move(Shogo + "/" +File_ary[i], item)
                  shutil.move(Shogo + "/" +File_ary[i + 1], item)
               i += 2
               
         else:
            # bool
            judge = False
            # エラー商号パス
            Error_Path = Shogo
            break

      
      return judge,Error_Path
   
   # 画像の二値化（画像パスとx座標,y座標,x2座標,y2座標）
   def Time_Parts_2_Ch_Img(self,path,x,y,x2,y2):
      # # Pillowで画像ファイルを開く
      pil_img = Match_Image.open(path)
      # PillowからNumPyへ変換
      img = np.array(pil_img)
      # RGBからBGRへ変換する
      if img.ndim == 3:
         im = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
      else:
         im = img
         
      # グレースケール化
      im_gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
      
      # 二値化(閾値を超えた画素を255にする。)
      threshold = 235# 閾値の設定
      ret, img_thresh = cv2.threshold(im_gray, threshold, 255, cv2.THRESH_BINARY)
      
      # 打刻データ部の始点(左上)x,y
      pt1 = (x, y)  
      # 打刻データ部の終点(右下)x,y
      pt2 = (x2, y2)
      
      # クロッピング変数
      cropped = img_thresh[pt2[0]:pt2[1],pt1[0]:pt1[1]]
      # 一時ファイルの作成
      fname = os.path.dirname(path) + '\\croppedimg.jpg'
      # 二値化画像の保存
      pil_image = Match_Image.fromarray(cropped)
      
      # Pillowで画像ファイルへ保存
      pil_image.save(fname)
      
      pil_cropped = Match_Image.open(fname)
      cropimg = np.array(pil_cropped)
      crpim = cv2.cvtColor(cropimg, cv2.COLOR_RGB2BGR)
      #print('cropimg.shape =',cropimg.shape)

      # 画像の合成
      fore_img = crpim
      back_img = im
      h, w = back_img.shape[:2]

      #合成用座標
      #合成用座標
      dx = pt1[0]      
      dy = pt2[0] 
      
      M = np.array([[1, 0, dx], [0, 1, dy]], dtype=float)
      img_warped = cv2.warpAffine(fore_img, M, (w, h), back_img, borderMode=cv2.BORDER_TRANSPARENT)
      #cv2.imwrite(expfilenm,img_warped)
      warped = cv2.cvtColor(img_warped, cv2.COLOR_BGR2RGB)
      pil_image = Match_Image.fromarray(warped)
      
      # Pillowで画像ファイルへ保存
      pil_image.save(path)

      # 一時データ削除
      os.remove(fname)
   
   def year_converter_to_wareki(self,date_obj):
      start_of_taisyo = date(1912, 7, 30) 
      start_of_syowa = date(1926, 12, 25) 
      start_of_heisei = date(1989, 1, 8) 
      start_of_reiwa = date(2019,5,1)
      year, month, day = None, None, None
      
      #-で分けてリスト化
      obj_list = date_obj.split("-")
      #2桁や1桁のときは19**にする
      if len(obj_list[0]) == 1:
         obj_list[0] = "190" + obj_list[0]
      elif len(obj_list[0]) == 2:
         obj_list[0] = "19" + obj_list[0]

      #年のみの入力は1月1日となる
      if len(obj_list) == 1:
         year = int(obj_list[0])
         input_obj = date(year, 1, 1)
      #年月の場合は1日となる
      elif len(obj_list) == 2:
         year, month = tuple(obj_list)
         year, month = int(year), int(month)
         input_obj = date(year, month, 1)
      elif len(obj_list) == 3:
         year, month, day = tuple(obj_list)
         year, month, day = int(year), int(month), int(day)
         input_obj = date(year, month, day)
      #リストに4つ以上の要素はエラー
      else:
         return None


      #出てきた年を格納する
      result_year = ""

      #年を格納
      if input_obj >=start_of_reiwa:
         nen = input_obj.year - start_of_reiwa.year + 1
         if nen == 1:
               nen = "元"
         result_year =  f"R{nen}"
      elif input_obj >= start_of_heisei:
         nen = input_obj.year - start_of_heisei.year + 1
         if nen == 1:
               nen = "元"
         result_year =  f"H{nen}"
      elif input_obj >= start_of_syowa:
         nen = input_obj.year - start_of_syowa.year + 1
         if nen == 1:
               nen = "元"
         result_year =  f"S{nen}"
      elif input_obj >= start_of_taisyo:
         nen = input_obj.year - start_of_taisyo.year + 1
         if nen == 1:
               nen = "元"
         result_year =  f"T{nen}"
      #大正よりも前の場合
      else:
         return "昔過ぎて計算できません（大正以前）"

      #output（年月日を合算）
      result = ""

      #年のみの入力の場合は年のみで返す
      if not month:
         result = result_year
      else:
         #年月の場合
         if not day:
               result = f"{result_year}{month}月"
         #年月日の場合
         else:
               result = f"{result_year}{month}月{day}日"
      return result
      
    # 処理対象フォルダ取得　 
   def Process_Folda_Create(self,thumnail_Folda_ary):
      # サムネイル配列
      Thum_ary = []
      # ﾀｲﾑｶｰﾄﾞ配列
      Timecard_ary = []   
      print(thumnail_Folda_ary)
      # サムネイルフォルダの配列ループ
      for thumnail_folda in glob.glob(thumnail_Folda_ary + "/*"):
         # サムネイルフォルダをソートしループ
         for process_fol in natsorted(glob.glob(thumnail_folda + "/*")):
            
            # ﾀｲﾑｶｰﾄﾞフォルダ
            if "TimeCard" in process_fol:
               # 列ﾌｧｲﾙパス取得
               for row_fol in natsorted(glob.glob(process_fol + "/*")):
                  Timecard_ary.append(row_fol)
            elif "Thumnail" in process_fol:
               # サムネイルフォルダ
               Thum_ary.append(process_fol)
      print(Thum_ary)
      return Timecard_ary,Thum_ary

   def ary_del_loop(self,result_ary,del_ary):
      del_cnt = 0
      for i in del_ary:
         del result_ary[i - del_cnt]
         del_cnt += 1
      return result_ary

class Excel_Operations():
   def __init__(self, Path):
       self.wb = load_workbook(Path)
       self.Result_Parh = Path
      #  # 現場リストパス
      #  self.genba_List_file_wb = load_workbook(shaingenba_path, read_only=True, keep_vba=True, data_only=True)
       # 現場の名前画像一時ファイルのパス
       self.Genba_Tmp_Img_path = "./asset/img/tmp/Genba_Img/Genba_temp.png"
       
   # シート作成関数
   def Sheet_Create(self):
      # シートの数の確認  
       sh_cnt = 0
       # デフォルトシートの確認  
       Def_Sh = False
       # 社員リストシート存在確認  
       Shain_List_bo = False
       # 社員現場詳細リスト シート存在確認
       Shain_Genba_Shosai_List = False
       
       # シート名の検索
       for ws in self.wb.worksheets:
         if ws.title == "社員現場詳細リスト":
               Shain_Genba_Shosai_List = True

       # 社員現場詳細リスト シート存在確認
       if Shain_Genba_Shosai_List == False:
           self.wb.create_sheet(index=--1, title="社員現場詳細リスト")
           shain_genba_sheet = self.wb["社員現場詳細リスト"]
           genbashosai_list = ["シート名","ﾀｲﾑｶｰﾄﾞ画像","主現場","氏名","副現場","氏名","新規","雇用区分","最終氏名"]
           
           i = 1
           for genba_item in genbashosai_list:
              # シートのヘッダ表記   
              shain_genba_sheet.cell(row=1, column=i).value = genba_item
              
              # シートのインデックスをローマ字化   
              cell_column_index = re.sub(r'[0-9]+', '',  shain_genba_sheet.cell(row=1,column=i).coordinate)
              
              # ﾀｲﾑｶｰﾄﾞの画像でセルの幅変化   
              if genba_item == "ﾀｲﾑｶｰﾄﾞ画像":
                  shain_genba_sheet.column_dimensions[cell_column_index].width = 65
              else:
                  shain_genba_sheet.column_dimensions[cell_column_index].width = 20
              
              i += 1 
              
       Last_Sheet_name = self.wb.worksheets[-1].title
       
       if Last_Sheet_name != "月間集計":
          sh_cnt = int(Last_Sheet_name) + 1
       else:
          sh_cnt = 1
         
       #  シートの数が最終シートのなる
       self.wb.copy_worksheet(self.wb.worksheets[0]).title = str(sh_cnt)
      
      #  カウントシートのアクティブ化
       return self.wb[str(sh_cnt)]
   
   # 社員現場リストの作成
   def Write_employee_list(self):
      # 社員リストシートをアクティブにするにする
      sheet = self.wb['月間集計']
      # 初期始まり行
      i = 4
      
      rtn_2nd_ary = []
      rtn_1st_ary = []
      
      # 締め日付取得
      Closing_Day = sheet["C2"].value
      
      while True:
      #   tmp_ary = []
        # 部門名取得
        Obtain_department_name = sheet.cell(row=i,column=3).value
        # 氏名取得
        Name = sheet.cell(row=i,column=4).value
        # 部門名が空であれば終了   
        if Obtain_department_name == "" or Obtain_department_name == None:
           break
         
        
        #  一次元配列の格納   
        rtn_1st_ary.append(Name)
        #  二次元配列の格納  
        rtn_2nd_ary.append(Obtain_department_name)
        
        i += 1
        
      return rtn_1st_ary,set(rtn_2nd_ary),Closing_Day
            
   # 社員現場リスト_一覧.xlsmから商号名を検索する
   def Sheet_Search(self,sheet,subject):
      # シートタイトルを格納する変数
      sheet_title = ""
      
      # シートの数ループ
      for sheet in self.genba_List_file_wb.worksheets:
         # 現場が見つかれば抜ける為のbool
         bool = False
         
         # 各シートの最終列までループ
         for cnt in range(1, sheet.max_column + 1):
            # 現場名の取得
            Get_val = sheet.cell(row=1, column=cnt).value

            # 引数が社員リストに含まれる場合ループを抜ける
            if subject in Get_val:
               # シート名の格納
               sheet_title = sheet.title
               bool == True
               
         # 現場名が見つかればループを抜ける
         if bool:
               break

      return sheet_title
   
   # 社員現場リスト_一覧.xlsmからシートのリストの作成
   def Employee_list_Create(self,sheet_name):
      #　シートを選択
      sheet = self.genba_List_file_wb[sheet_name]
      # 社員現場リスト
      Employee_list = []
      # 列ループ
      for col in range(1,sheet.max_column + 1):
         # 始点行
         rows = 1
         # 一時配列
         tmp_list = []
         # 行が空になるまでループ
         while True:
             # 行に記載されている値を取得
             val = sheet.cell(row=rows,column=col).value
             #  行に記載されている値が空の時にループ抜ける
             #  値が空でなければ一時配列にあたいを格納
             if val != None:
               tmp_list.append(val) 
             else:
               break
            # 次の行へ 
             rows += 1
         
         # 社員現場リストに一時配列を格納 
         Employee_list.append(tmp_list)
      
      return Employee_list
   
    #  リストの書き込み
   def List_Writes(self,sheet,List,pos):
      # リスト変数
      List_val = ""
      if type(List)==str:
            List_val = List
            # リスト作成
            dv = DataValidation(
                  type="list",
                  formula1=List_val,
            )
      else:
         # 取得した名前リストが空でなければ
         if len(List) != 0:
            # 名前リストのループ
            for item in List:
               List_val += item + ","
         
      
         # リスト作成
         dv = DataValidation(
               type="list",
               formula1="\""+List_val+"\"",
         )

      # 指定セル番地に書き込み
      dv.add(pos)

      # シートに追加する
      sheet.add_data_validation(dv)
            
   def daterange(self, _start, _end):
        for n in range((_end - _start).days):
            yield _start + timedelta(n)
            
   def Days_List_Create(self,shori_date):
      # 日付ライブラリの定義
      dt = datetime.datetime.today()
      
      Dt_year = dt.year
      # 処理月
      Process_month = shori_date.month
      
      # 日付テストの値
      # Process_month = 3
      
      # 処理日
      Process_day = shori_date.day
      
      if Process_month == 1:
         # デフォルト計算始まり年（今年度）
         st_year = str(Dt_year - 1)
         # デフォルト計算始まり月（今年度）
         st_month = "12"
      else:
         # デフォルト計算始まり年（今年度）
         st_year = str(Dt_year)
         st_month = str(Process_month)
      
      # デフォルト計算始まり年（今年度）
      ed_year = str(Dt_year)

      # デフォルト計算終わり月（今年度）
      ed_month = str(Process_month + 1)

      
      # 0埋め
      if len(st_month) == 1:
        st_month = "0" + st_month
      if len(ed_month) == 1:
        ed_month = "0" + ed_month
      
      # 処理月の前月の最終日を取得
      last_month_Final_day = self.get_last_date(datetime.date(int(st_year), int(st_month), int(Process_day))).day
      
      # 計算はじめのyyyymmdd
      st = st_year + st_month + str(int(Process_day))
      
      # 計算終わり日付
      ed = ed_year + ed_month + str(int(Process_day))
      
      # 日付変換（計算始まり変数）
      start = datetime.datetime.strptime(st, '%Y%m%d')
      # 日付変換（計算終わり変数）
      end = datetime.datetime.strptime(ed, '%Y%m%d')
      
      rtn_yyyymmdd_list = []
      
      # 日付の記載
      for i in self.daterange(start, end):
         # 返り値配列
         rtn_yyyymmdd_list.append(i)   
         
         # 処理対象日が最終日と同じでTRUE
         if i.day == last_month_Final_day:
             # 31日合わせの為の配列 
             for item in range(0,31 - last_month_Final_day):
                 rtn_yyyymmdd_list.append("")
      
      return rtn_yyyymmdd_list
   
   def get_last_date(self,dt):
       return dt.replace(day=calendar.monthrange(dt.year, dt.month)[1])

   def isBizDay(self,DATE):
      holiday = jpholiday.is_holiday_name(DATE)
      holiday_cnt = 0
      week_day = DATE.weekday()
      Sunday_cnt = 0
      
      Saturday_cnt = 0
      
      # 祝日の時のカウント
      if holiday != None:
         holiday_cnt += 1
      
      # 土曜日の時のカウント
      if week_day == 5:
         Saturday_cnt += 1
      # 日曜日の時のカウント
      if week_day == 6:
         Sunday_cnt += 1
         
      return Saturday_cnt,Sunday_cnt,holiday_cnt
              
   def Days_Border(self, sheet, kyokai_row):
      last_cell_No = self.Last_Row_Search(1, 4, sheet)
      # 青
      Blue_Line = Side(style='medium', color='4682b4')
      # 赤
      Red_Line = Side(style='medium', color='dc143c')
      # 対象最終列
      taisho_col = 8
      
      # 青線定義（セルの描画位置は下線）
      border_lr_blue = Border(bottom=Blue_Line)
      # 赤線定義（セルの描画位置は下線）
      border_lr_red = Border(top=Red_Line)
      # 中間線
      for Col in range(1,taisho_col):
         # 上線の描画
         sheet.cell(row=3, column=Col).border = border_lr_blue
         # 中間線の描画
         sheet.cell(row=kyokai_row, column=Col).border = border_lr_blue
         # 下線の描画
         sheet.cell(row=last_cell_No, column=Col).border = border_lr_red
      
      # 範囲線の描画
      # 青線（セルの描画位置は左線）
      border_lr_blue = Border(left=Blue_Line)
      # 赤線（セルの描画位置は左線）
      border_lr_red = Border(left=Red_Line)
      # 行ループ（始点は、4行目）
      for Row in range(4,last_cell_No):
         if Row <= kyokai_row:
            sheet.cell(row=Row, column=taisho_col).border = border_lr_blue
         else: 
            sheet.cell(row=Row, column=taisho_col).border = border_lr_red
              
   def Excel_Img_Past(self,sheet,img_ary,pos):
      print(img_ary)
      for thumnail in img_ary:
         img = Image(thumnail)
         #リサイズ
         img.width = 500
         img.height = 600
         sheet.add_image(img, pos)
         
         # タイムカード画像の貼り付け
         im = Match_Image.open(thumnail)
         # 名前画像の切り取り
         im_crop = im.crop((0, 0, im.width, 200))
         # 画像の保存
         im_crop.save(self.Genba_Tmp_Img_path, quality=100)

   def Shain_Shosai_Create(self, Sheet_Cnt, Shozoku_Name,Name_List,Bumon_List):
      # 社員現場リストをアクティブにする
      Genba_sheet = self.wb["社員現場詳細リスト"]
      Last_Row = self.Last_Row_Search(1, 1, Genba_sheet)
                  
      # 行の高さ 
      Genba_sheet.row_dimensions[Last_Row].height = 80
      
      # シート名の記載
      Genba_sheet.cell(row=Last_Row,column=1).value = Sheet_Cnt
      
      # ﾀｲﾑｶｰﾄﾞ画像の添付
      # 貼り付け画像の選定
      img = Image(self.Genba_Tmp_Img_path)
      # 画像の横幅
      img.width = 450
      # 画像の高さ
      img.height = 100
      
      Last_Row_index = str(Last_Row)
      
      # 画像を添付
      Genba_sheet.add_image(img, "B" + Last_Row_index)
      
      Genba_sheet.cell(row=Last_Row,column=3).value = Shozoku_Name
      self.List_Writes(Genba_sheet, Name_List, "D" + Last_Row_index)
      self.List_Writes(Genba_sheet, Bumon_List, "E" + Last_Row_index)
      self.List_Writes(Genba_sheet, Name_List, "F" + Last_Row_index)
      self.List_Writes(Genba_sheet, Bumon_List, "H" + Last_Row_index)
      
      Genba_sheet.cell(row=Last_Row,column=9).value = "=IF(D" + Last_Row_index + " <> \"\",D"+ Last_Row_index + ",IF(F"+ Last_Row_index + "<> \"\","+ Last_Row_index + ",IF(G" + Last_Row_index + " <> \"\",G" + Last_Row_index + ",\"\")))"
      # 各シートのリンク
      Genba_sheet["A" + Last_Row_index].hyperlink = os.path.basename(self.Result_Parh) + "#" + str(Last_Row - 1) + "!A1"
      Genba_sheet["A" + Last_Row_index].font = Font(size=20, bold=True,color='0067c0')
      # Genba_sheet["A" + Last_Row_index].font = openpyxl.styles.fonts.Font(color='0067c0')
      Genba_sheet["A" + Last_Row_index].alignment = Alignment(horizontal="centerContinuous",vertical="center")
 
   # 列の最終行調べる
   def Last_Row_Search(self, st_col, st_row,sheet):
        l = st_row
        Last_Row = 1
        while True:
            if sheet.cell(row=l, column=st_col).value == None:
                Last_Row = l
                break
            l += 1

        return Last_Row
   
   # Excelの保存
   def Excel_Save(self):
      self.wb.save(self.Result_Parh)
   
   def Shozoku_Member_List(self,Path,shozoku):
      genba_List_wb = load_workbook(
            Path, read_only=True, keep_vba=True, data_only=True)
      Name_List = []
        
      # 商号名   
      shogo_Name_rtn = ""

      # シートの数ループ
      for sheet in genba_List_wb.worksheets:
         # 現場が見つかれば抜ける為のbool
         genba_j = False
     
         # 各シートの最終列までループ
         for genba_name_cnt in range(1, sheet.max_column + 1):
               # 現場名の取得
               genba_name = sheet.cell(row=1, column=genba_name_cnt).value

            #  Shogo_Genba_List += genba_name +
               # ｼｽﾃﾑ引数が現場名に含まれているか判定
               if shozoku in genba_name:
                  # 社員リストの作成
                  shogo_Name_rtn,Name_List,genba_j = self.Shain_List_Create(2,genba_name_cnt,sheet,True)
                           
         # 現場名が見つかればループを抜ける
         if genba_j:
               break
      
      
      if shogo_Name_rtn != "":
         Shogo_Sheet = genba_List_wb[shogo_Name_rtn]
         # genba_List_wb.active = genba_List_wb.worksheets[shogo_Name_rtn]
         genba_list =[]
         
         for shogo_sheet_cnt in range(1, Shogo_Sheet.max_column):
             genba_list.append(self.Shain_List_Create(1,shogo_sheet_cnt,Shogo_Sheet,False))
         
                
      return shogo_Name_rtn,Name_List,genba_list
   
   def Shain_List_Create(self,rows,col,sheet,rt_judge):
      genba_j = False
      shogo_Name_rtn = ""
      Name_List = []
      # 名前リスト作成
      # セルが空になるまでループ
      while True:
         # 社員名の取得
         shain_Name = sheet.cell(
               row=rows, column=col).value
         
         # 空行の場合ループを抜ける
         if shain_Name == None:
               # 外部ループ抜ける判定
               genba_j = True
               shogo_Name_rtn = sheet.title
               break
         else:
               # 名前リストに名前を追加
               Name_List.append(shain_Name)
               # 嗣の行へ
               rows += 1
      if rt_judge: 
         return shogo_Name_rtn,Name_List,genba_j
      else:
         return Name_List
      
   def border_create(self, color, syurui):
        border = Border(
            left=Side(
                border_style=syurui,
                color=color,

            ),
            right=Side(
                border_style=syurui,
                color=color
            ),
            top=Side(
                border_style=syurui,
                color=color
            ),
            bottom=Side(
                border_style=syurui,
                color=color

            )
        )

        return border
     
   def Merge_Cells(self,sheet,hani):
      # セルの範囲結合
      sheet.merge_cells(hani)
      # 特定位置を検索
      idx = hani.find(":")
      # 特定位置より前の文字列の取得
      cell_pos = hani[:idx]
      # セル中央設定
      center_alignment = Alignment(horizontal='center',vertical='center')
      # セルのセンタリング
      sheet[cell_pos].alignment = center_alignment
   
   def Sheet_Chk(self,sheet):
      # 処理開始列
      st_col = 2
      # 処理終了列
      ed_col = 6
      # 処理開始行
      st_row = 4
      # 最終行取得
      ed_row =  self.Last_Row_Search(1, 4,sheet)

      # for col in range(st_col,ed_col):
      for rows in range(st_row, ed_row):
         # 2列目の値
         row_ary = []

         # 3列目の値
         for col in range(st_col, ed_col):
            # 行列の値の格納
            row_ary.append(sheet.cell(row=rows, column=col).value)
         
         # 2,3,4,5列全てが空ではなく4列埋まっている時
         if row_ary.count("") != 4 and row_ary.count("") != 0:
            # 2列目が空のときは全てエラー
            if row_ary[0] == "":
               self.Er_mark(sheet,row_ary, rows)
            else:
               # 3列目が空
               if row_ary[1] == "": 
                  # 4列目が空
                  if row_ary[2] == "":
                     # 5列目が空
                     if row_ary[3] == "":
                        # 空列に対しエラーマーク処理を行う
                        self.Er_mark(sheet,row_ary, rows)
                  else:
                     # 空列に対しエラーマーク処理を行う
                     self.Er_mark(sheet,row_ary, rows)
               else:
                  if int(row_ary[0]) > int(row_ary[1]):
                     self.Er_val_Mark(sheet,  rows, 2)
                     self.Er_val_Mark(sheet,  rows, 3)
                  
                  # 4列目が空
                  if row_ary[2] == "":
                     # 5列目が空
                     if row_ary[3] != "":
                        # 空列に対しエラーマーク処理を行う
                        self.Er_mark(sheet,row_ary, rows)
                  else:
                     if int(row_ary[1]) > int(row_ary[2]):
                        self.Er_val_Mark(sheet,  rows, 3)
                        self.Er_val_Mark(sheet,  rows, 4)
                        
                     # 5列目が空
                     if row_ary[3] == "":
                        # 空列に対しエラーマーク処理を行う
                        self.Er_mark(sheet, row_ary, rows)
                     
                     else:
                        if int(row_ary[2]) > int(row_ary[3]):
                           self.Er_val_Mark(sheet,  rows, 4)
                           self.Er_val_Mark(sheet,  rows, 5)
         
         col = 2
         
         for item in row_ary:
             if item != "" and item != None:
               if int(item) > 2400:
                 self.Er_val_Mark(sheet,  rows, col)
             col += 1
                              
   def Er_mark(self,sheet, row_ary,rows):
      # 空列のインデックス
      None_index = 2
      # 赤線定義
      Red_Line = Side(style='medium', color='dc143c')
      # ボーダー位置（上下左右）
      border_lr_red = Border(top=Red_Line, bottom=Red_Line,
                             left=Red_Line, right=Red_Line,)
      # 行ループ
      for item in row_ary:
         # 値が空の時に赤線描画
         if item == "":
            sheet.cell(row=rows, column=None_index).border = border_lr_red
         None_index += 1
   
   def Er_val_Mark(self,sheet,row,col):
      # 赤線定義
      Red_Line = Side(style='medium', color='dc143c')
      # ボーダー位置（上下左右）
      border_lr_red = Border(top=Red_Line, bottom=Red_Line,
                             left=Red_Line, right=Red_Line,)

      sheet.cell(row=row, column=col).border = border_lr_red
   
   def Sheet_Ocr_framework_create(self,sheet,month,result_row_ary,yyymmdd):
      
      # シート位置
      sh_cnt = str(int(sheet.title) + 1)
      
      # 列の結合
      for i in range(3,7):
            self.Merge_Cells(sheet,"B" + str(i) + ":" + "C" + str(i))
            
      self.Merge_Cells(sheet,"B2:C2")
      self.Merge_Cells(sheet,"D2:E2")
      sheet["B2"] = "社員現場詳細リンク"
      sheet["B2"].hyperlink = os.path.basename(self.Result_Parh) + "#社員現場詳細リスト!A" + str(sh_cnt)
      sheet["B2"].font = openpyxl.styles.fonts.Font(color='0067c0')
      sheet["D2"] = "月間集計"
      sheet["D2"].hyperlink = os.path.basename(self.Result_Parh) + "#月間集計!A1"
      sheet["D2"].font = openpyxl.styles.fonts.Font(color='0067c0')
      
      # 基本情報の入力
      sheet["A2"] = str(yyymmdd.month) + "月"   
      sheet["A3"] = "主現場"
      sheet["B3"] = "=IF(社員現場詳細リスト!C" + sh_cnt + "<>\"\",社員現場詳細リスト!C" + sh_cnt + ",\"\")"
      sheet["D3"] = "=IF(社員現場詳細リスト!D" + sh_cnt + "<>\"\",社員現場詳細リスト!D" + sh_cnt + ",\"\")"
      sheet["A4"] = "副現場"
      sheet["B4"] = "=IF(社員現場詳細リスト!E" + sh_cnt + "<>\"\",社員現場詳細リスト!E" + sh_cnt + ",\"\")"
      sheet["D4"] = "=IF(社員現場詳細リスト!F" + sh_cnt + "<>\"\",社員現場詳細リスト!F" + sh_cnt + ",\"\")"
      # 新規行処理
      sheet["A5"] = "新規"
      sheet["D5"] = "=IF(社員現場詳細リスト!G" + sh_cnt + "<>\"\",社員現場詳細リスト!G" + sh_cnt + ",\"\")"
      # 雇用区分行処理
      sheet["A6"] = "雇用区分"
      sheet["D6"] =  "=IF(社員現場詳細リスト!H" + sh_cnt + "<>\"\",社員現場詳細リスト!H" + sh_cnt + ",\"\")"
      
      # 反映先の勤怠欄の名前
      sheet["AD3"] =  "=IF(社員現場詳細リスト!H" + sh_cnt + "<>\"\",社員現場詳細リスト!H" + sh_cnt + ",\"\")"
      # 反映先の勤怠欄の雇用区分
      sheet["Z3"] =  "=IF(社員現場詳細リスト!I" + sh_cnt + "<>\"\",社員現場詳細リスト!I" + sh_cnt + ",\"\")"
      
      # 結果の集計始点行
      row = 7
      # 結果の集計始点列
      
      main = []
      for row_ary in result_row_ary:
         col = 1
         c = 5
         tmp_ary = []
         for col_val in row_ary:
            sheet.cell(row=row,column=col).value = col_val
            if col == 1:
               sheet.cell(row=row,column=col).number_format = "m/d"
            else:
               sheet.cell(row=row,column=col).number_format = "##\":\"##"
               Col_Number = utils.cell.get_column_letter(col) + str(row)
               sheet.cell(row=row,column=col + c).value = "=IF(LEN("+ Col_Number +") = 3,LEFT(" + Col_Number + ",1),IF(LEN(" + Col_Number + ") = 4,LEFT(" + Col_Number + ",2),\"\"))"
               sheet.cell(row=row,column=col + (c + 1)).value = "=IF(AND(LEN(" + Col_Number + ") >= 3, LEN(" + Col_Number + ") <= 4),RIGHT(" + Col_Number + ",2),\"\")"
               tmp_ary.append(utils.cell.get_column_letter(col + c) + str(row))
               tmp_ary.append(utils.cell.get_column_letter(col + (c + 1)) + str(row))
               c += 1
            col += 1
         
         w_col = 27
         for row_val in tmp_ary:
            sheet.cell(row=row,column=w_col).value = "=IF(" + str(row_val) + " <> \"\"," + str(row_val) + ",0)"
            w_col += 1
            
         row += 1
      
      # 非表示の処理
      for hid in tmp_ary:
         newstring = re.sub(r'[0-9]+', '', hid)
         sheet.column_dimensions[newstring].hidden= True